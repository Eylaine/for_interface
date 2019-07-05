# coding=utf-8
# Time   : 2019/6/13 10:27
# Author : zhonglin.zhang
# File   : excel.py

from openpyxl import Workbook, load_workbook
from settins import ROOT_PATH


class ExcelRead:

    def __init__(self, file_name):
        self.file_name = ROOT_PATH + file_name
        self.work_book = load_workbook(self.file_name)
        self.sheet_name = 'Sheet1'
        self.sheet = self.work_book[self.sheet_name]

    def set_sheet(self, sheet_name):
        self.sheet_name = sheet_name
        self.sheet = self.work_book[sheet_name]

    def get_by_index(self, row_index, col_index):
        """index: 从1开始，不是从0开始"""
        return self.sheet.cell(row_index, col_index).value

    def get_by_location(self, location):
        """location, egs: A1, B2 etc."""
        return self.sheet[location].value
